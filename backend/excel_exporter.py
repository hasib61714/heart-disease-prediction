from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from typing import List
from app.database import PredictionHistory, PatientProfile
from sqlalchemy.orm import Session

def create_patients_excel(db: Session) -> BytesIO:
    """Create Excel file with all patient predictions"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Predictions"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID", "Patient ID", "Patient Name", "Age", "Gender", "Phone",
        "Chest Pain Type", "Blood Pressure", "Cholesterol", "Heart Rate",
        "Prediction", "Risk %", "Doctor Notes", "Date"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Get data with patient profiles
    predictions = db.query(PredictionHistory).join(PatientProfile).all()
    
    # Data rows
    for row_idx, pred in enumerate(predictions, 2):
        profile = pred.profile
        
        data = [
            pred.id,
            profile.patient_id,
            profile.name,
            pred.age,
            "Male" if pred.sex == 1 else "Female",
            profile.phone,
            get_chest_pain_text(pred.cp),
            pred.trestbps,
            pred.chol,
            pred.thalach,
            pred.prediction,
            f"{pred.risk_probability:.2f}%",
            pred.doctor_notes or "",
            pred.created_at.strftime("%Y-%m-%d %H:%M")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Color code by risk
            if col == 11:  # Prediction column
                if value == "High Risk":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                else:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def get_chest_pain_text(cp: int) -> str:
    """Convert chest pain code to text"""
    cp_map = {
        0: "Typical Angina",
        1: "Atypical Angina",
        2: "Non-anginal Pain",
        3: "Asymptomatic"
    }
    return cp_map.get(cp, "Unknown")


def create_high_risk_patients_excel(db: Session) -> BytesIO:
    """Create Excel file with only high-risk patients"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "High Risk Patients"
    
    # Header style
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Patient ID", "Name", "Phone", "Age", "Gender",
        "Risk %", "BP", "Cholesterol", "Latest Checkup", "Doctor Notes"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Get high-risk predictions
    predictions = db.query(PredictionHistory)\
        .join(PatientProfile)\
        .filter(PredictionHistory.prediction == "High Risk")\
        .order_by(PredictionHistory.risk_probability.desc())\
        .all()
    
    # Data rows
    for row_idx, pred in enumerate(predictions, 2):
        profile = pred.profile
        
        data = [
            profile.patient_id,
            profile.name,
            profile.phone,
            pred.age,
            "Male" if pred.sex == 1 else "Female",
            f"{pred.risk_probability:.2f}%",
            pred.trestbps,
            pred.chol,
            pred.created_at.strftime("%Y-%m-%d"),
            pred.doctor_notes or ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Highlight high risk percentage
            if col == 6:
                risk_val = float(value.strip('%'))
                if risk_val > 80:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif risk_val > 60:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
